#!/usr/bin/env python3
"""make_coded_reference.py -- build the HISTORICAL reading (`D`) that the real-figure
validation scores against.

Primary source: the per-article extraction workbooks
  <DRIVE>/thesis/meta-analysis/DISSERTATION/rodent/RODENT_Processed_Extractions/*.xlsm
sheet 'Main', a key-value table whose columns C.. are Comparison_1..N.

These are strictly richer than the final meta-analysis CSV and are the right source:

  * ~198 articles instead of the 43 that survived the MA's CONTENT screen -- an article
    excluded for the wrong population or outcome still has a perfectly valid figure to
    read, so the extraction-validation universe is much larger than the MA's.
  * per-arm `*_Variance_Type` (SD / SEM / CI-95), recorded explicitly rather than
    inferred from convention. This is the single most consequential field in the study:
    reading SEM as SD multiplies the SD by sqrt(n).
  * `Control_Group_ID` / `Experimental_Group_ID` -- the arm NAMES (e.g. SC / EE), which
    give a real reference for series->arm binding, the silent catastrophic class.
  * `Direction`, `Multi_Arm_Study`, `Shared_Control_Group`,
    `Variance_Inflation_Factor`, `Between_Or_Within_Design` -- everything the
    end-to-end metafor stage needs.

Falls back to the final CSV when the workbooks are unavailable.

Writes coded/coded_reference.json in the schema of ANALYSIS-PLAN.md sec.10.3.
Adds nothing it cannot source: `controlLandmarkId`/`intervLandmarkId` are left null for
the annotator to fill, because guessing which bar is which arm would fabricate exactly
the mis-assignment the study exists to measure.

  python3 make_coded_reference.py               # build
  python3 make_coded_reference.py --stats       # population table only, no write
"""
import argparse
import collections
import json
import math
import os
import pathlib
import re
import sys

HERE = pathlib.Path(__file__).resolve().parent
OUT = HERE / "coded"

# The hand-coded dissertation data lives in a separate, private repository. Resolution order
# matches benchmark/real/inventory.py so the two agree: an explicit env var, then that repo
# cloned as a SIBLING of this one, then the author's own absolute paths as a last resort.
# (The absolute defaults used to be the only option, which made this script unrunnable by
# anyone else and contradicted the README's promise of RODENT_CSV / HUMAN_CSV overrides.)
_SIBLING = HERE.parent.parent.parent / "GSF-dissertation-meta-analysis" / "data" / "raw"
WORKBOOKS = pathlib.Path(os.environ.get(
    "RODENT_WORKBOOKS",
    "/mnt/c/Users/gregs/My Drive/thesis/meta-analysis/DISSERTATION/rodent/"
    "RODENT_Processed_Extractions"))
CSV_FALLBACK = pathlib.Path(os.environ.get("RODENT_CSV")
                            or (_SIBLING / "rodent_data.csv" if (_SIBLING / "rodent_data.csv").exists()
                                else "/mnt/c/Users/gregs/GSF-dissertation-meta-analysis/"
                                     "data/raw/rodent_data.csv"))

FIG_RE = re.compile(r"^\s*fig(?:ure)?\.?\s*([0-9]+)\s*[-_ ]?\s*([A-Za-z])?", re.I)
# Data_Extraction_Method strings that CLAIM the number is also stated in text/table.
# "Claim" is the operative word -- see verify_oracle_v2.py; this flag is a CANDIDATE marker
# and never by itself admits a row to the accuracy analysis.
ORACLE_CLAIM = ("text/figure", "text and figure", "text and figures",
                "figure/table", "inset table", "open-access data")


def parse_data_source(ds):
    """'Figure 5B' -> (figure '5', panel 'B'). Returns (None, None) when not a figure."""
    if not ds:
        return (None, None)
    m = FIG_RE.match(str(ds).strip())
    if not m:
        return (None, None)
    return (m.group(1), (m.group(2) or "").upper() or None)


def is_figure_derived(ds, dm):
    ds, dm = str(ds or "").strip().lower(), str(dm or "").strip().lower()
    if ds.startswith("fig"):
        return True
    return any(k in dm for k in ("from figure", "figure analysis", "in figure",
                                 "re-extracted", "text/figure", "text and figure",
                                 "visual analysis"))


def rounding_quantum_pct(text_value):
    s = str(text_value).strip()
    try:
        v = float(s)
    except (TypeError, ValueError):
        return None
    if v == 0:
        return None
    dp = len(s.split(".")[1]) if "." in s else 0
    return round(100 * (0.5 * 10 ** (-dp)) / abs(v), 4)


def to_sd(value, vtype, n):
    """Per-arm variance -> SD, using the RECORDED type rather than a convention."""
    if value is None or n in (None, 0):
        return None
    t = str(vtype or "").upper().replace("-", "").replace(" ", "")
    try:
        v, n = float(value), float(n)
    except (TypeError, ValueError):
        return None
    if t == "SD":
        return v
    if t in ("SEM", "SE"):
        return v * math.sqrt(n)
    if t in ("CI95",):
        return (v / 1.96) * math.sqrt(n)
    if t in ("CI99",):
        return (v / 2.576) * math.sqrt(n)
    return None                              # unknown type -> refuse, do not guess


def read_workbooks(root):
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl not installed -- falling back to the CSV", file=sys.stderr)
        return None
    import warnings
    warnings.filterwarnings("ignore")
    files = sorted(root.glob("*.xlsm"))
    if not files:
        return None
    rows = []
    for fp in files:
        try:
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            if "Main" not in wb.sheetnames:
                continue
            ws = wb["Main"]
            grid = [list(r) for r in ws.iter_rows(values_only=True)]
        except Exception as e:
            print(f"  [warn] {fp.name}: {e}", file=sys.stderr)
            continue
        hdr = next((i for i, r in enumerate(grid)
                    if r and str(r[0]).strip() == "Category"), None)
        if hdr is None:
            continue
        ncomp = sum(1 for c in grid[hdr][2:] if c)
        fields = {}
        for r in grid[hdr + 1:]:
            if not r or len(r) < 3 or r[1] is None:
                continue
            fields[str(r[1]).strip()] = list(r[2:2 + ncomp])
        for j in range(ncomp):
            def g(name):
                v = fields.get(name, [None] * ncomp)
                v = v[j] if j < len(v) else None
                return None if v is None or str(v).strip() == "" else v
            rows.append({k: g(k) for k in (
                "Article_ID", "DOI", "Comparison_ID", "Data_Source",
                "Data_Extraction_Method", "Direction", "Control_N", "Intervention_N",
                "Control_Group_Mean", "Control_Group_Variance_Value",
                "Control_Group_Variance_Type", "Intervention_Group_Mean",
                "Intervention_Group_Variance_Value", "Intervention_Group_Variance_Type",
                "Control_Group_ID", "Experimental_Group_ID", "Multi_Arm_Study",
                "Number_of_Arms", "Shared_Control_Group", "Variance_Inflation_Factor",
                "Between_Or_Within_Design", "Test_Name", "Variable_Name",
                "Cognitive_Domain")} | {"_file": fp.name})
    return rows


def read_csv_fallback(path):
    import csv as _csv
    if not path.exists():
        return []
    out = []
    for r in _csv.DictReader(open(path)):
        out.append({
            "Article_ID": r.get("Article_ID"), "DOI": r.get("DOI"),
            "Comparison_ID": r.get("Comparison_ID"), "Data_Source": r.get("Data_Source"),
            "Data_Extraction_Method": r.get("Data_Extraction_Method"),
            "Direction": r.get("Direction"),
            "Control_N": r.get("Control_N"), "Intervention_N": r.get("Intervention_N"),
            "Control_Group_Mean": r.get("Control_Group_Mean"),
            "Control_Group_Variance_Value": r.get("Control_Group_Variance_Value"),
            "Control_Group_Variance_Type": None,          # not in the final CSV
            "Intervention_Group_Mean": r.get("Intervention_Group_Mean"),
            "Intervention_Group_Variance_Value": r.get("Intervention_Group_Variance_Value"),
            "Intervention_Group_Variance_Type": None,
            "Control_Group_ID": r.get("Control_Group_ID_Standardized"),
            "Experimental_Group_ID": r.get("Intervention_Group_ID_Standardized"),
            "Multi_Arm_Study": r.get("Multi_Arm_Study"),
            "Number_of_Arms": r.get("Number_of_Arms"),
            "Shared_Control_Group": r.get("Shared_Control_Group"),
            "Variance_Inflation_Factor": r.get("VIF_multiarm"),
            "Between_Or_Within_Design": None,
            "Test_Name": r.get("Test_Name"), "Variable_Name": r.get("Variable_Name"),
            "Cognitive_Domain": r.get("Cognitive_Domain"), "_file": "rodent_data.csv",
        })
    return out


def build(rows):
    out = []
    for r in rows:
        ds, dm = r.get("Data_Source"), r.get("Data_Extraction_Method")
        if not is_figure_derived(ds, dm):
            continue
        fignum, letter = parse_data_source(ds)
        art = str(r.get("Article_ID") or "").strip()
        if not art:
            continue
        cn, inn = r.get("Control_N"), r.get("Intervention_N")
        cvt = r.get("Control_Group_Variance_Type")
        ivt = r.get("Intervention_Group_Variance_Type")
        rec = {
            "article": art, "doi": r.get("DOI"),
            # `comparisonId` is the workbook's OWN Comparison_ID, kept for traceability back to
            # the source -- it is NOT unique (one id can cover several panels, e.g. Zhang2017_3_1
            # is both "Figure 3d" and "Figure 3e"). Anything that keys on it silently collapses
            # rows. Use `rowId` as the join key; it is made unique below.
            "comparisonId": str(r.get("Comparison_ID") or f"{art}_?"),
            "rowId": None,                # assigned after the loop, guaranteed unique
            "dataSource": ds, "extractionMethod": dm,
            "figureNumber": fignum, "panelLetter": letter,
            "figureId": f"{art}_fig{fignum}" if fignum else None,
            "figureDerived": True,
            "oracleClaimed": any(k in str(dm or "").lower() for k in ORACLE_CLAIM),
            "oracleVerified": None,       # filled ONLY by verify_oracle_v2.py
            "isOracle": False,            # never true until verification says so
            "direction": r.get("Direction"),
            "design": r.get("Between_Or_Within_Design"),
            "testName": r.get("Test_Name"), "variableName": r.get("Variable_Name"),
            "domain": r.get("Cognitive_Domain"),
            "controlArmName": r.get("Control_Group_ID"),
            "intervArmName": r.get("Experimental_Group_ID"),
            "multiArm": str(r.get("Multi_Arm_Study")).strip().lower() == "true",
            "nArms": r.get("Number_of_Arms"),
            "sharedControl": str(r.get("Shared_Control_Group")).strip().lower() == "true",
            "vifMultiarm": _f(r.get("Variance_Inflation_Factor"), 1.0),
            "control": _arm(r.get("Control_Group_Mean"),
                            r.get("Control_Group_Variance_Value"), cvt, cn),
            "interv": _arm(r.get("Intervention_Group_Mean"),
                           r.get("Intervention_Group_Variance_Value"), ivt, inn),
            "controlLandmarkId": None, "intervLandmarkId": None,
            "_source": r.get("_file"),
        }
        rec["complete"] = all(rec[k].get("sd") is not None and rec[k].get("mean") is not None
                              and rec[k].get("n") for k in ("control", "interv"))
        out.append(rec)
    _assign_row_ids(out)
    return out


def _assign_row_ids(recs):
    """Give every record a unique, stable, human-traceable `rowId`.

    Base is article::comparisonId::dataSource -- which disambiguates the common case (one
    Comparison_ID spanning several panels). Where even that collides (the same panel yielding
    several outcomes), a #n ordinal is appended in encounter order. Emits a warning naming the
    source collisions so the upstream data can be fixed rather than silently worked around."""
    import collections
    src_dupes = collections.Counter(r["comparisonId"] for r in recs)
    seen = collections.Counter()
    for r in recs:
        base = f'{r["article"]}::{r["comparisonId"]}::{str(r.get("dataSource") or "?").strip().lower()}'
        seen[base] += 1
        r["rowId"] = base if seen[base] == 1 else f"{base}#{seen[base]}"
    ids = [r["rowId"] for r in recs]
    assert len(ids) == len(set(ids)), "rowId must be unique"
    collided = {k: v for k, v in src_dupes.items() if v > 1}
    if collided:
        n = sum(collided.values())
        print(f"  [warn] {len(collided)} source Comparison_ID values are reused across {n} rows "
              f"(e.g. {sorted(collided)[0]!r}); `comparisonId` is NOT a key -- use `rowId`.")


def _f(v, default=None):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _arm(mean, var, vtype, n):
    return {
        "mean": _f(mean), "n": _f(n),
        "varianceValue": _f(var), "varianceType": (str(vtype).strip() if vtype else None),
        "sd": to_sd(var, vtype, n),
        "roundingQuantumPct": rounding_quantum_pct(var),
        "meanRoundingQuantumPct": rounding_quantum_pct(mean),
    }


def stats(recs):
    arts = {r["article"] for r in recs}
    panels = {(r["article"], r["dataSource"]) for r in recs}
    figs = {(r["article"], r["figureNumber"]) for r in recs if r["figureNumber"]}
    vt = collections.Counter()
    for r in recs:
        for k in ("control", "interv"):
            vt[r[k].get("varianceType") or "MISSING"] += 1
    q = sorted(r[k]["roundingQuantumPct"] for r in recs for k in ("control", "interv")
               if r[k].get("roundingQuantumPct") is not None)
    print(f"  comparisons (figure-derived) : {len(recs)}")
    print(f"  articles                     : {len(arts)}")
    print(f"  distinct (article, panel)    : {len(panels)}")
    print(f"  distinct (article, figure)   : {len(figs)}")
    print(f"  complete mean+SD+n both arms : {sum(1 for r in recs if r['complete'])}")
    print(f"  ORACLE-CLAIMED (unverified)  : {sum(1 for r in recs if r['oracleClaimed'])}"
          f"  across {len({r['article'] for r in recs if r['oracleClaimed']})} articles")
    print(f"  named panel letter           : {sum(1 for r in recs if r['panelLetter'])}")
    print(f"  multi-arm / shared control   : {sum(1 for r in recs if r['multiArm'])}"
          f" / {sum(1 for r in recs if r['sharedControl'])}")
    print(f"  arm-value readings (2/comp)  : {2*len(recs)}")
    print("  per-arm VARIANCE TYPE        : "
          + ", ".join(f"{k}={v}" for k, v in vt.most_common()))
    if q:
        print(f"  coded rounding quantum %     : median {q[len(q)//2]:.3f}"
              f"  p90 {q[int(.9*len(q))]:.3f}  >1%: {sum(1 for x in q if x > 1)}/{len(q)}")
    dm = collections.Counter(str(r["extractionMethod"]) for r in recs)
    print("  Data_Extraction_Method       :")
    for k, v in dm.most_common(12):
        print(f"      {v:>4}  {k}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--stats", action="store_true", help="print the population, do not write")
    ap.add_argument("--workbooks", default=str(WORKBOOKS))
    a = ap.parse_args()

    rows = read_workbooks(pathlib.Path(a.workbooks))
    src = "workbooks"
    if not rows:
        print(f"  workbooks unavailable at {a.workbooks} -- using the final CSV")
        rows = read_csv_fallback(CSV_FALLBACK)
        src = "rodent_data.csv"
    if not rows:
        sys.exit("no coded source available")
    recs = build(rows)
    print(f"CODED REFERENCE (source: {src})")
    stats(recs)
    if a.stats:
        return
    OUT.mkdir(exist_ok=True)
    (OUT / "coded_reference.json").write_text(json.dumps(
        {"source": src, "nRows": len(recs), "rows": recs}, indent=2, default=str))
    print(f"\n[written] {OUT/'coded_reference.json'}")
    print("  NOTE: isOracle is FALSE for every row until verify_oracle_v2.py confirms the "
          "value is actually printed in the paper's text. See ANALYSIS-PLAN sec.1.4b.")


if __name__ == "__main__":
    main()
